import sqlite3
from os import system, getcwd, path, popen, listdir, name, chdir
from os.path import join

from pandas import DataFrame, read_sql_query, to_datetime, concat,ExcelWriter, Series
from datetime import datetime , time
from numpy import nan, max, min, sum , mean, int64, array_split, log, nanmean, deg2rad, rad2deg, array, isnan
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.cm as cm
from   openpyxl import Workbook
from   openpyxl import load_workbook 
from   collections import defaultdict
from   statistics import mean
from scipy.stats import circmean

import math

class DB:    
    
    def __init__(self) -> None:
        self.sqlite3            = sqlite3
        self.system             = system
        self.getcwd             = getcwd
        self.path               = path
        self.listdir            = listdir
        self.name               = name
        self.chdir              = chdir
        self.join               = join

        # PANDAS
        self.DataFrame          = DataFrame
        self.read_sql_query     = read_sql_query
        self.datetime           = datetime
        self.to_datetime        = to_datetime
        self.concat             = concat
        self.ExcelWriter        = ExcelWriter
        self.Series             = Series

        # NUMPY
        self.nan                = nan
        self.max                = max
        self.min                = min
        self.sum                = sum
        self.mean               = mean
        self.int64              = int64
        self.array_split        = array_split
        self.log                = log
        self.nanmean            = nanmean
        self.deg2rad            = deg2rad
        self.rad2deg            = rad2deg
        self.array              = array
        self.isnan              = isnan

        # SCIPY
        self.circmean           = circmean


        self.time               = time        
        self.plt                = plt
        self.mdates             = mdates
        self.cm                 = cm
        self.Workbook           = Workbook
        self.load_workbook      = load_workbook
        self.defaultdict        = defaultdict
        self.math               = math
        self.con                = None
        self.dbName             = ""
        self.saveToFilePath     = self.join(self.getcwd(),"output",'MorganLewis.xlsx') 
        self.ops                = {"max": self.max,"min":self.min, "mean":self.mean, "nanmean":self.nanmean,"circmean": self.circularMean,"sum":self.sum}

         

    def connect(self,name) -> sqlite3.Connection:
        '''RETURN A CONNECTION TO SQLITE DB'''
        print("CONNECTING....")
        # print(self.getcwd())
        # print(self.listdir())
        # self.chdir()
        dir = self.join(self.getcwd(),"database")
        path = self.join(self.getcwd(),"database",name)
        print(path)
        if name in self.listdir(dir):
            self.dbName = name
            try:
                con = self.sqlite3.connect(path)
            except Exception as e:
                print(f"Error connecting to {name}  => {str(e)}") 
            else:
                self.con = con
                return True
        else:
            return False
        

    def test(self):
        '''TEST CONNECTION SQLITE DB'''
        if self.con:
            print(f"Connected to : {self.dbName}") 
            cur = self.con.cursor()
            res = cur.execute('SELECT * FROM Remote_Data WHERE IDRemote="2009" ORDER BY "Timestamp" DESC LIMIT 10')
            print(res.fetchone())
        else:
            print(f"Unable to connect to {self.dbName}")

    def execute_query(self, query):
            '''Execute SQLITE3 QUERY'''
            cursor = self.con.cursor()
            try:
                    cursor.execute(query)
                    self.con.commit()
                    print("Query executed successfully")
            except self.Error as e:
                    print(f"The error '{e}' occurred")


    def getDataset(self,variable,unit):
        '''PULL ALL THE DATA FOR A SPECIFIC VARIABLE FROM SQLITE DB'''
        # AT, RH, BARO, DP, AT2, QNH, RAIN, WDI, WSI, GUST, GUSTDIR, QFE, QFF, WDA, WSA, 
        # BATTERY, IR, UV, VIS, SLP, SP, ATMAX, ATMIN, SOLRAD, DAYRAIN, SOLRAD_ACC, HOURS_OF_SUN,
        # CFB, PLS, RLS, RAIN10M, WL, BV, WL2, ITEMP, BVS, COND, TURB, DO, WTEMP, pH, RAIN05M, LEVEL, LEVEL2, SAL, WL1, QNH_inHg
      
        sq 	        = self.read_sql_query(f"SELECT RecordStamp,Reading FROM Remote_Data WHERE IDRemote='2009' AND Sense='{variable}' ORDER BY 'RecordStamp' DESC",self.con,parse_dates={"RecordStamp":"%Y-%m-%d %H:%M"})
        sq.columns  = ['Date',f"{variable}({unit})"]
        sq.index    = sq['Date']
        sq 		    = sq.drop(columns=['Date'])
        sq          = sq.loc[ ~sq.index.duplicated(), :] # REMOVE DUPLICATES
        return sq

    def getDatasets(self,getDataFor):
        '''PULL ALL REQUIRED DATA FROM SQLITE DB'''
        # AT, RH, BARO, DP, AT2, QNH, RAIN, WDI, WSI, GUST, GUSTDIR, QFE, QFF, WDA, WSA, 
        # BATTERY, IR, UV, VIS, SLP, SP, ATMAX, ATMIN, SOLRAD, DAYRAIN, SOLRAD_ACC, HOURS_OF_SUN,
        # CFB, PLS, RLS, RAIN10M, WL, BV, WL2, ITEMP, BVS, COND, TURB, DO, WTEMP, pH, RAIN05M, LEVEL, LEVEL2, SAL, WL1, QNH_inHg

        frames = []
        for variable in getDataFor.keys(): 
            sq 	        = self.read_sql_query(f"SELECT RecordStamp,Reading FROM Remote_Data WHERE IDRemote='2009' AND Sense='{variable}' ORDER BY 'RecordStamp' DESC",self.con,parse_dates={"RecordStamp":"%Y-%m-%d %H:%M"})
            sq.columns  = ['Date',f"{variable}({getDataFor[variable]})"]
            sq.index    = sq['Date']
            sq 		    = sq.drop(columns=['Date'])
            sq          = sq.loc[ ~sq.index.duplicated(), :] # REMOVE DUPLICATES 
            frames.append(sq)
            
        # print("CONCATENATING ALL DATAFRAMES")    
        gt = self.concat(frames,axis=1,verify_integrity=True) 
        
        # gt.where(((gt < 361) & (gt > 1)), self.nan, inplace=True) # REMOVE AMBIGUOUS VALUES
        # with self.ExcelWriter("raw.xlsx") as writer:    
        #     gt.to_excel(writer,index= True,sheet_name = f"RAW" )

        return gt
    
    def circularMean(self,x):
        angles = self.deg2rad(x.values[~self.isnan(x.values)]) 
        circmean = self.circmean(angles)        
        return self.rad2deg(circmean)
 

    def circularMeanTest(self,x):
        angles = self.deg2rad(self.array(x))
        circmean = self.circmean(angles)        
        return self.rad2deg(circmean)
    
 

    def dailyAvg(self,df,variables):
         '''RETURN  / SAVE TO EXCEL FILE THE DAILY AVG FOR SELECT VARIABLES'''
         query  = ["Year","Month","Day","Hour","Min"] 
         query.extend([f"{x['param']}({x['unit']})"  for x in variables])  # ["ATMAX(℃)","ATMIN(℃)","Year","Month","Day"]

         #  AGGREGATOR
         aggregator     = {}
         for v in variables:
              aggregator[f"{v['param']}({v['unit']})"] =  self.ops[v["func"]]

         sq             = df.copy()
         sq['Year']     = sq.index.year
         sq['Month']    = sq.index.month
         sq['Day']      = sq.index.day
         sq['Hour']     = sq.index.hour
         sq['Min']      = sq.index.minute
         #  daily          = sq[query].groupby(["Year","Month","Day"]).agg(aggregator).round(1)

         with self.ExcelWriter(self.saveToFilePath,mode='w') as writer:   
            sq[query].to_excel(writer,index= False,sheet_name = f"10 Min Avg") 
            # daily.to_excel(writer,index= True,sheet_name = f"Daily Avg")

         return sq

     
    def TemperatureMaxAndMin(self, variables ):
        '''RETURN  / SAVE TO EXCEL FILE THE DAILY MAX & MIN FOR SELECT VARIABLES'''
        '''sq,{"ATMAX(℃)":"max","ATMIN(℃)":"min"}'''

        query  = ["Year","Month","Day"]
        query.extend([f"{x['param']}({x['unit']})"  for x in variables])  # ["ATMAX(℃)","ATMIN(℃)","Year","Month","Day"]

        #  AGGREGATOR
        aggregator     = {}
        for v in variables:
            aggregator[f"{v['param']}({v['unit']})"] =  self.ops[v["func"]]

        #  PARAMS
        params = {}
        for v in variables:
            params[v['param']] =  v['unit']

        sq             = self.getDatasets(params) # {'ATMAX':'℃','ATMIN':'℃'}
        sq.where(((sq <= 40) & (sq >= 0)), self.nan, inplace=True) # REMOVE AMBIGUOUS VALUES. VALUES > 50 ℃ /  10min SINCE AT IS 10 MINUTELY
        #  print(sq.head())
        sq['Year']     = sq.index.year
        sq['Month']    = sq.index.month
        sq['Day']      = sq.index.day
        daily          = sq[query].groupby(["Year","Month","Day"]).agg(aggregator)
        daily.reset_index(inplace=True)

        with self.ExcelWriter(self.saveToFilePath,mode='a') as writer:    
            daily.to_excel(writer,index= False,sheet_name = f"Daily Max & Min Air Temp" )

        return daily


    def Temperature(self,param):
        '''RETURN  / SAVE TO EXCEL FILE THE MONTHLY AVG TEMPERATURE '''
        
        label   = f"{param['variable']}({param['unit']})"
        query   = ["Year","Month","Day",label]

        sq                 = self.getDataset(param['variable'],param['unit'])
        sq.where(((sq <= 40) & (sq >= 0)), self.nan, inplace=True) # REMOVE AMBIGUOUS VALUES. VALUES > 50 ℃ /  10min SINCE AT IS 10 MINUTELY

        sq['Year']         = sq.index.year
        sq['Month']        = sq.index.month
        sq['Day']          = sq.index.day

        daily              = sq[query].groupby(["Year","Month","Day"]).agg({label: self.nanmean})        
        # max_days_indexes   = daily.groupby(["Year","Month"])[label].idxmax()  # GET INDEX FOR ALL MAX AT DAYS
        # maxDays            = daily.loc[max_days_indexes]
        # monthly            = daily.groupby(["Year","Month"]).agg({label:self.mean}).round(1)

        # ADJUST HEADERS
        daily.columns       = ["AT(℃)"]
        # maxDays.columns     = ["AT(℃)"]
        # monthly.columns     = ["AT(℃)"]

        # RESET INDEX
        # maxDays.reset_index(inplace=True)
        # monthly.reset_index(inplace=True)

        with self.ExcelWriter(self.saveToFilePath,mode='w') as writer:  
            sq[["Year","Month","Day",label]].to_excel(writer,index= False,sheet_name = f"AT" )
            # daily.to_excel(writer,index= True,sheet_name = f"Daily AT" )    

        return sq[label]

    def TemperatureMax(self,param):
        '''RETURN  / SAVE TO EXCEL FILE MAX TEMPERATURE '''
        
        label   = f"{param['variable']}({param['unit']})"
        query   = ["Year","Month","Day",label]

        sq                 = self.getDataset(param['variable'],param['unit'])
        sq.where(((sq <= 40) & (sq >= 0)), self.nan, inplace=True) # REMOVE AMBIGUOUS VALUES. VALUES > 50 ℃ /  10min SINCE ATMAX IS 10 MINUTELY

        sq['Year']         = sq.index.year
        sq['Month']        = sq.index.month
        sq['Day']          = sq.index.day

        daily              = sq[query].groupby(["Year","Month","Day"]).agg({label: self.max})       
        # max_days_indexes   = daily.groupby(["Year","Month"])[label].idxmax()  # GET INDEX FOR ALL MAX ATMAX DAYS
        # maxDays            = daily.loc[max_days_indexes]
        # monthly            = daily.groupby(["Year","Month"]).agg({label:self.mean}).round(1)

        # ADJUST HEADERS
        daily.columns       = ["ATMAX(℃)"]
        # maxDays.columns     = ["ATMAX(℃)"]
        # monthly.columns     = ["ATMAX(℃)"]

        # RESET INDEX
        # maxDays.reset_index(inplace=True)
        # monthly.reset_index(inplace=True)

        with self.ExcelWriter(self.saveToFilePath,mode='a') as writer:   
            sq[["Year","Month","Day",label]].to_excel(writer,index= False,sheet_name = f"ATMAX(RAW)" ) 
        return sq[label]

    def TemperatureMin(self,param):
        '''RETURN  / SAVE TO EXCEL FILE MIN TEMPERATURE '''
        
        label   = f"{param['variable']}({param['unit']})"
        query   = ["Year","Month","Day",label]

        sq                 = self.getDataset(param['variable'],param['unit'])
        sq.where(((sq <= 40) & (sq >= 0)), self.nan, inplace=True) # REMOVE AMBIGUOUS VALUES. VALUES > 50 ℃ /  10min SINCE ATMIN IS 10 MINUTELY

        sq['Year']         = sq.index.year
        sq['Month']        = sq.index.month
        sq['Day']          = sq.index.day

        daily              = sq[query].groupby(["Year","Month","Day"]).agg({label: self.max})       
        # max_days_indexes   = daily.groupby(["Year","Month"])[label].idxmax()  # GET INDEX FOR ALL MAX ATMIN DAYS
        # maxDays            = daily.loc[max_days_indexes]
        # monthly            = daily.groupby(["Year","Month"]).agg({label:self.mean}).round(1)

        # ADJUST HEADERS
        daily.columns       = ["ATMIN(℃)"]
        # maxDays.columns     = ["ATMIN(℃)"]
        # monthly.columns     = ["ATMIN(℃)"]

        # RESET INDEX
        # maxDays.reset_index(inplace=True)
        # monthly.reset_index(inplace=True)

        with self.ExcelWriter(self.saveToFilePath,mode='a') as writer: 
        # with self.ExcelWriter('TEMP.xlsx',mode='w') as writer:  
            sq[["Year","Month","Day",label]].to_excel(writer,index= False,sheet_name = f"ATMIN(RAW)") 
        return sq[label]

    def WindSpeed(self,param):
        '''RETURN  / SAVE TO EXCEL FILE FOR WINDSPEED '''
        
        label   = f"{param['variable']}({param['unit']})"
        query   = ["Year","Month","Day","Hour","Min",label]

        sq                 = self.getDataset(param['variable'],param['unit'])
        sq.where(((sq <= 40) & (sq >= 0)), self.nan, inplace=True) # REMOVE AMBIGUOUS VALUES. VALUES > 40 KT /  10min SINCE WINDSPEED IS 10 MINUTELY

        sq['Year']      = sq.index.year
        sq['Month']     = sq.index.month
        sq['Day']       = sq.index.day
        sq['Hour']      = sq.index.hour
        sq['Min']       = sq.index.minute

        daily           = sq[query].groupby(["Year","Month","Day"]).agg({label: self.nanmean})        
        # max_days_indexes   = daily.groupby(["Year","Month"])[label].idxmax()  # GET INDEX FOR ALL MAX WINDSPEED DAYS
        # maxDays            = daily.loc[max_days_indexes]
        # monthly            = daily.groupby(["Year","Month"]).agg({label:self.mean}).round(1)

        # ADJUST HEADERS
        daily.columns       = ["WSA(KT)"]
        # maxDays.columns     = ["WSA(KT)"]
        # monthly.columns     = ["WSA(KT)"]

        # RESET INDEX
        # maxDays.reset_index(inplace=True)
        # monthly.reset_index(inplace=True)


        with self.ExcelWriter(self.saveToFilePath,mode='a') as writer:  
            sq[query].to_excel(writer,index= False,sheet_name = f"WindSpeed" )
         
        return sq[label]

    def WindDirection(self,param):
        '''RETURN  / SAVE TO EXCEL FILE FOR WIND DIRECTION '''
        
        label   = f"{param['variable']}({param['unit']})"
        query   = ["Year","Month","Day","Hour","Min",label]

        sq                 = self.getDataset(param['variable'],param['unit'])
        sq.where(((sq <= 360) & (sq >= 0)), self.nan, inplace=True) # REMOVE AMBIGUOUS VALUES. VALUES > 360° || < 0 

        sq['Year']      = sq.index.year
        sq['Month']     = sq.index.month
        sq['Day']       = sq.index.day
        sq['Hour']      = sq.index.hour
        sq['Min']       = sq.index.minute

        daily           = sq[query].groupby(["Year","Month","Day"]).agg({label: self.circularMean})        
        # max_days_indexes   = daily.groupby(["Year","Month"])[label].idxmax()  # GET INDEX FOR ALL MAX DIRECTION DAYS
        # maxDays            = daily.loc[max_days_indexes]
        # monthly            = daily.groupby(["Year","Month"]).agg({label:self.mean}).round(1)

        # ADJUST HEADERS
        daily.columns       = ["WDA(°)"]
        # maxDays.columns     = ["WDA(°)"]
        # monthly.columns     = ["WDA(°)"]

        # RESET INDEX
        # maxDays.reset_index(inplace=True)
        # monthly.reset_index(inplace=True)

        # with self.ExcelWriter('windDir.xlsx',mode='w') as writer: 
        with self.ExcelWriter(self.saveToFilePath,mode='a') as writer:  
            sq[query].to_excel(writer,index= False,sheet_name = f"WindDir" )
            # daily[[label]].to_excel(writer,index= True,sheet_name = f"WDA DAILY" )
         
        return sq[label]

    def Rainfall(self,param):
        '''RETURN  / SAVE TO EXCEL FILE THE MONTHLY AVG RAINFALL, MAX RAINFALL DAY FOR EACH MONTH'''
        
        label   = f"{param['variable']}({param['unit']})"
        query   = ["Year","Month","Day",label]

        sq                 = self.getDataset(param['variable'],param['unit'])
        sq.where(((sq <= 20) & (sq >= 0)), self.nan, inplace=True) # REMOVE AMBIGUOUS VALUES. VALUES > 20mm/5min SINCE RAIN RESOLUTION IS 5 MINUTELY

        sq['Year']         = sq.index.year
        sq['Month']        = sq.index.month
        sq['Day']          = sq.index.day

        daily              = sq[query].groupby(["Year","Month","Day"]).agg({label: self.sum})         
        max_days_indexes   = daily.groupby(["Year","Month"])[label].idxmax()  # GET INDEX FOR ALL MAX RAINFALL DAYS
        maxDays            = daily.loc[max_days_indexes]
        monthly            = daily.groupby(["Year","Month"]).agg({label:self.mean}).round(1)

        # ADJUST HEADERS
        daily.columns       = ["RAIN(mm/day)"]
        maxDays.columns     = ["RAIN(mm/day)"]
        monthly.columns     = ["RAIN(mm/month)"]

        # RESET INDEX
        maxDays.reset_index(inplace=True)
        monthly.reset_index(inplace=True)


        with self.ExcelWriter(self.saveToFilePath,mode='a') as writer:    
            # sq[label].to_excel(writer,index= True,sheet_name = f"Rainfall(RAW)" ) #
            # daily.to_excel(writer,index= True,sheet_name = f"Daily Rainfall Total" ) #
            maxDays.to_excel(writer,index= False,sheet_name = f"Day With Max Rainfall")
            monthly.to_excel(writer,index= False,sheet_name = f"Average Monthly Rainfall")
        return sq[label]
         
         
    def RelativeHumidity(self,param):
        '''RETURN  / SAVE TO EXCEL FILE THE MONTHLY AVG RH'''

        label           = f"{param['variable']}({param['unit']})"        

        sq              = self.getDataset(param['variable'],param['unit'])
        sq.where(((sq <= 100) & (sq > 0)), self.nan, inplace=True) # REMOVE AMBIGUOUS VALUES

        sq['Year']      = sq.index.year
        sq['Month']     = sq.index.month 
        sq['Day']       = sq.index.day

        monthly         = sq[["Year","Month",label]].groupby(["Year","Month" ]).agg({label: self.nanmean}).round(0)
        dailyMaxAndMin  = sq[["Year","Month","Day",label]].groupby(["Year","Month","Day"]).agg(Max=(label,self.max),Min=(label,self.min)).round(0)

        # RESET INDEX
        dailyMaxAndMin.reset_index(inplace=True)
        monthly.reset_index(inplace=True)

        with self.ExcelWriter(self.saveToFilePath,mode='a') as writer:    
            monthly.to_excel(writer,index= False,sheet_name = f"Average Monthly RH" )
            dailyMaxAndMin.to_excel(writer,index= False,sheet_name = f"RH Daily Max & Min" ) 
        return sq[label]
    

    def DownTime(self,frames):
        sq = self.concat(frames,axis=1,verify_integrity=True) 
                 
        sq['Year']      = sq.index.year
        sq['Month']     = sq.index.month 
        sq['Day']       = sq.index.day
        daily           = sq.groupby(["Year","Month","Day"]).agg({'AT(℃)': self.nanmean,'WSA(KT)':self.nanmean,'WDA(°)':self.circularMean,'RAIN(mm)':self.mean,'RH(%)':self.nanmean}).round(1)
        daily['DownDays']   = daily.apply(lambda x: False in self.isnan(x.values),axis=1) 
        gt              = daily[(daily.DownDays == False) ]  
        DT              = gt['DownDays'].groupby(["Year","Month"]).count()
        
        
    

        with self.ExcelWriter(self.saveToFilePath,mode='a') as writer:    
            # sq.to_excel(writer,index= True,sheet_name = f"COMBINED" )
            # daily.to_excel(writer,index= True,sheet_name = f"C DAILY" ) 
            DT.to_excel(writer,index= True,sheet_name = f"Down Time" ) 


             


if __name__ == '__main__':
     db = DB()

     a = [249, 204, nan, 131, 326, 251, 132, 224, 239, 249, 218, 189, 201, 121, 148, 175, 146, 145, 130, 134, 128, 139, 128, 124, 131, 170, 137, 161, 141, 133, 119, 63, 136, 134, 124, 117, 115, 148, 63, 119, 124, 125, 119, 119, 153, 131, 130, 63, 129, 134, 130, 117, 138, 113, 113, 111, 63, 118, 119, 116, 122, 132, 121, 124, 116, 107, 120, 111, 100, 105, 116, 108, 115, 110, 115, 109, 88, 114, 140, 130, 109, 108, 109, 119, 138, 143, 125, 122, 104, 104]
     print(db.circularMeanTest(a))